from django.contrib import admin
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from .models import User, UserProfile, EmailVerificationToken, PhoneVerificationToken
import openpyxl
from django.http import HttpResponse
from datetime import datetime



def export_users_to_excel(modeladmin, request, queryset):
    """
    Export selected users to an Excel file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    # Header row
    ws.append(["Full Name", "Email", "Phone", "Account Type", "Date Joined"])

    for user in queryset:
        ws.append([
            user.get_full_name(),
            user.email,
            user.phone,
            user.account_type,
            user.date_joined.strftime("%Y-%m-%d %H:%M"),
        ])

    # Create HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = f"users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename={filename}'
    wb.save(response)
    return response

export_users_to_excel.short_description = "📤 Export selected users to Excel"


def export_emails(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Emails"
    ws.append(["Email"])
    for user in queryset:
        ws.append([user.email])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="emails.xlsx"'
    wb.save(response)
    return response
export_emails.short_description = "📧 Export Emails Only"


def export_phones(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Phones"
    ws.append(["Phone"])
    for user in queryset:
        ws.append([user.phone])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="phones.xlsx"'
    wb.save(response)
    return response
export_phones.short_description = "📱 Export Phones Only"


class UserProfileInline(admin.StackedInline):
    model = UserProfile
    can_delete = False
    readonly_fields = ('tasks_completed', 'tasks_posted', 'success_rate', 'average_rating', 'total_reviews')
    fieldsets = (
        ('Profile Info', {
            'fields': ('occupation', 'company', 'location', 'website', 'skills', 'experience_years')
        }),
        ('Social Links', {
            'fields': ('twitter_url', 'linkedin_url', 'facebook_url')
        }),
        ('Stats', {
            'fields': ('tasks_completed', 'tasks_posted', 'success_rate', 'average_rating', 'total_reviews')
        }),
    )


@admin.register(User)
class UserAdmin(BaseUserAdmin):
    actions = [export_users_to_excel, export_emails, export_phones]    
    list_display = ('email', 'get_full_name', 'account_type', 'is_active', 'email_verified', 'phone_verified', 'date_joined')
    list_filter = ('account_type', 'is_active', 'email_verified', 'phone_verified', 'date_joined')
    search_fields = ('email', 'first_name', 'last_name', 'phone')
    ordering = ('-date_joined',)

    fieldsets = (
        (None, {'fields': ('email', 'password')}),
        ('Personal info', {'fields': ('first_name', 'last_name', 'phone', 'preferred_currency')}),
        ('Role & Permissions', {'fields': ('account_type', 'is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions')}),
        ('Verification', {'fields': ('email_verified', 'phone_verified')}),
        ('Important dates', {'fields': ('last_login', 'date_joined')}),
    )

    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': ('email', 'password1', 'password2', 'account_type'),
        }),
    )

    readonly_fields = ('date_joined', 'last_login')

    inlines = [UserProfileInline]   # Attach profile inline


@admin.register(UserProfile)
class UserProfileAdmin(admin.ModelAdmin):
    actions = [export_users_to_excel, export_emails, export_phones]

    list_display = ('user', 'occupation', 'experience_years', 'tasks_completed', 'tasks_posted', 'success_rate', 'average_rating')
    list_filter = ('experience_years', 'created_at')
    search_fields = ('user__email', 'user__first_name', 'user__last_name', 'occupation', 'company')
    readonly_fields = ('tasks_completed', 'tasks_posted', 'success_rate', 'average_rating', 'total_reviews')


@admin.register(EmailVerificationToken)
class EmailVerificationTokenAdmin(admin.ModelAdmin):
    list_display = ('user', 'token', 'created_at', 'expires_at', 'used')
    list_filter = ('used', 'created_at', 'expires_at')
    search_fields = ('user__email', 'token')
    readonly_fields = ('token',)


@admin.register(PhoneVerificationToken)
class PhoneVerificationTokenAdmin(admin.ModelAdmin):
    list_display = ('user', 'token', 'created_at', 'expires_at', 'used')
    list_filter = ('used', 'created_at', 'expires_at')
    search_fields = ('user__phone', 'token')
    readonly_fields = ('token',)
